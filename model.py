import requests
import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import numpy as np
from statsmodels.tsa.arima.model import ARIMA
from datetime import datetime
import streamlit as st

API_KEY = st.secrets.get("API_KEY", "your_actual_fmp_api_key")  # Use Streamlit secrets or replace with your key
SAVE_PATH = os.path.expanduser("~/Desktop/Financial_Models")
os.makedirs(SAVE_PATH, exist_ok=True)

# === CLEAN LABELS ===
def clean_label(s):
    """Cleans and formats financial statement labels."""
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s).replace('_', ' ').title()
    replacements = {
        'Ebitda': 'EBITDA', 'Ebit': 'EBIT', 'Opex': 'Operating Expenses', 'R D': 'R&D',
        'Cogs': 'Cost of Goods Sold', 'Capex': 'Capital Expenditures',
        'Netincome': 'Net Income', 'Depreciationandamortization': 'Depreciation & Amortization',
        'Deferredincometax': 'Deferred Income Tax', 'Deferredrevenue': 'Deferred Revenue',
        'Totalcurrentassets': 'Total Current Assets', 'Totalcurrentliabilities': 'Total Current Liabilities',
        'Totalstockholdersequity': 'Total Stockholders Equity', 'Totalassets': 'Total Assets',
        'Totaldebt': 'Total Debt', 'Shortterminvestments': 'Short Term Investments',
        'Propertyplantequipmentnet': 'Property Plant & Equipment Net',
        'Cashandcashequivalents': 'Cash & Cash Equivalents',
        'Operatingcashflow': 'Operating Cash Flow', 'Freecashflow': 'Free Cash Flow',
        'Interestexpense': 'Interest Expense', 'Grossprofit': 'Gross Profit',
        'Operatingincome': 'Operating Income', 'Revenue': 'Revenue',
        'Researchanddevelopmentexpenses': 'Research & Development Expenses'
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s.strip()

# === FETCH DATA ===
def fetch_fmp_data(statement, ticker, years, api_key):
    """Fetches financial data from FMP API."""
    url = f"https://financialmodelingprep.com/api/v3/{statement}/{ticker}?limit={years}&apikey={api_key}"
    try:
        r = requests.get(url)
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"Error fetching {statement} for {ticker}: {e}")
        return pd.DataFrame()

    data = r.json()
    if not isinstance(data, list) or not data:
        print(f"No data found for {statement} for {ticker}.")
        return pd.DataFrame()

    df = pd.DataFrame(data)
    drop_cols = ['symbol', 'cik', 'reportedCurrency', 'fillingDate', 'acceptedDate',
                 'calendarYear', 'period', 'link', 'finalLink']
    df.drop(columns=[col for col in drop_cols if col in df.columns], inplace=True)

    df['date'] = pd.to_datetime(df['date'])
    df.set_index("date", inplace=True)
    df.sort_index(inplace=True)
    df = df.transpose()
    df.columns = [col.strftime('%Y') for col in df.columns]
    df.index = [clean_label(i) for i in df.index]
    df = df.apply(pd.to_numeric, errors='coerce') / 1_000_000
    return df.round(1)

# === FETCH MARKET DATA ===
def fetch_market_data(ticker, api_key):
    """Fetches market data for EPS and P/E ratio."""
    url = f"https://financialmodelingprep.com/api/v3/quote/{ticker}?apikey={api_key}"
    try:
        r = requests.get(url)
        r.raise_for_status()
        data = r.json()
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame([{
            'date': d['priceDate'],
            'eps': d.get('eps', np.nan),
            'pe': d.get('pe', np.nan)
        } for d in data])
        df['date'] = pd.to_datetime(df['date'])
        df.set_index('date', inplace=True)
        df.sort_index(inplace=True)
        df = df.transpose()
        df.columns = [col.strftime('%Y') for col in df.columns]
        return df
    except requests.RequestException as e:
        print(f"Error fetching market data for {ticker}: {e}")
        return pd.DataFrame()

# === GROWTH INDEX ===
def growth_index(df):
    """Calculates growth index and CAGR."""
    if df.empty:
        return pd.DataFrame(), pd.Series()

    df_sorted = df.sort_index(axis=1)
    growth_df = pd.DataFrame(index=df_sorted.index, columns=df_sorted.columns, dtype=float)
    cagr_series = pd.Series(index=df_sorted.index, dtype=float)

    for index, row in df_sorted.iterrows():
        valid = row.dropna()
        if valid.empty or valid.iloc[0] == 0:
            growth_df.loc[index] = np.nan
            cagr_series[index] = np.nan
            continue
        growth_df.loc[index] = (row / valid.iloc[0]).round(2)
        if len(valid) >= 2:
            cagr = ((valid.iloc[-1] / valid.iloc[0]) ** (1 / (len(valid) - 1)) - 1) * 100
            cagr_series[index] = round(cagr, 2)
        else:
            cagr_series[index] = np.nan

    return growth_df, cagr_series

# === ANOMALY DETECTION ===
def detect_anomalies(df):
    """Detects anomalies using IQR method."""
    anomalies = {}
    for index, row in df.iterrows():
        valid = row.dropna()
        if len(valid) < 3:
            continue
        q1, q3 = valid.quantile([0.25, 0.75])
        iqr = q3 - q1
        outliers = valid[(valid < q1 - 1.5 * iqr) | (valid > q3 + 1.5 * iqr)]
        if not outliers.empty:
            anomalies[index] = [f"Anomaly in {year}: {value}M (outside IQR [{q1:.1f}, {q3:.1f}])" for year, value in outliers.items()]
    return anomalies

# === FORECASTING ===
def forecast_metric(series, steps=2):
    """Forecasts a metric using ARIMA."""
    try:
        valid = series.dropna()
        if len(valid) < 3:
            return pd.Series([np.nan] * steps, index=[str(int(valid.index[-1]) + i) for i in range(1, steps + 1)])
        model = ARIMA(valid, order=(1, 1, 1)).fit()
        forecast = model.forecast(steps)
        return pd.Series(forecast, index=[str(int(valid.index[-1]) + i) for i in range(1, steps + 1)]).round(1)
    except Exception as e:
        print(f"Forecasting error for {series.name}: {e}")
        return pd.Series([np.nan] * steps, index=[str(int(series.index[-1]) + i) for i in range(1, steps + 1)])

# === SENTIMENT ANALYSIS ===
def analyze_sentiment(ticker):
    """Placeholder for sentiment analysis using X API and Grok 4."""
    return f"Sentiment for {ticker}: Positive (based on recent X posts)"  # Placeholder

# === FINANCIAL RATIOS ===
def calculate_ratios(is_df, bs_df, cf_df, market_df):
    """Calculates financial ratios."""
    common_years = sorted(list(set(is_df.columns) & set(bs_df.columns) & set(cf_df.columns) & set(market_df.columns)))
    if not common_years:
        print("No common years for ratio calculation.")
        return pd.DataFrame()

    is_df_filtered = is_df[common_years]
    bs_df_filtered = bs_df[common_years]
    cf_df_filtered = cf_df[common_years]
    market_df_filtered = market_df[common_years]

    def get_row(df, key):
        return df.loc[key] if key in df.index else pd.Series([np.nan] * len(df.columns), index=df.columns)

    def safe_div(numerator, denominator):
        denominator_numeric = pd.to_numeric(denominator, errors='coerce')
        return numerator / denominator_numeric.replace(0, np.nan)

    ratios = pd.DataFrame(index=common_years)

    revenue = get_row(is_df_filtered, "Revenue")
    gross_profit = get_row(is_df_filtered, "Gross Profit")
    operating_income = get_row(is_df_filtered, "Operating Income")
    ebitda = get_row(is_df_filtered, "EBITDA")
    net_income = get_row(is_df_filtered, "Net Income")
    ebit = get_row(is_df_filtered, "EBIT")
    interest_expense = get_row(is_df_filtered, "Interest Expense")
    eps = get_row(market_df_filtered, "eps")
    pe = get_row(market_df_filtered, "pe")

    total_assets = get_row(bs_df_filtered, "Total Assets")
    total_stockholders_equity = get_row(bs_df_filtered, "Total Stockholders Equity")
    total_current_assets = get_row(bs_df_filtered, "Total Current Assets")
    total_current_liabilities = get_row(bs_df_filtered, "Total Current Liabilities")
    inventory = get_row(bs_df_filtered, "Inventory")
    cash_and_equivalents = get_row(bs_df_filtered, "Cash & Cash Equivalents")
    short_term_investments = get_row(bs_df_filtered, "Short Term Investments")
    total_debt = get_row(bs_df_filtered, "Total Debt")
    debt_for_ratio = total_debt if not total_debt.isnull().all() else get_row(bs_df_filtered, "Total Liabilities")
    property_plant_equipment_net = get_row(bs_df_filtered, "Property Plant & Equipment Net")

    free_cash_flow = get_row(cf_df_filtered, "Free Cash Flow")
    operating_cash_flow = get_row(cf_df_filtered, "Operating Cash Flow")
    capital_expenditure = get_row(cf_df_filtered, "Capital Expenditure")

    ratios["Gross Margin"] = safe_div(gross_profit, revenue)
    ratios["Operating Margin"] = safe_div(operating_income, revenue)
    ratios["EBITDA Margin"] = safe_div(ebitda, revenue)
    ratios["Net Profit Margin"] = safe_div(net_income, revenue)
    ratios["Gross Profit Ratio"] = safe_div(gross_profit, revenue)
    ratios["Operating Income Ratio"] = safe_div(operating_income, revenue)
    ratios["Return on Assets (ROA)"] = safe_div(net_income, total_assets)
    ratios["Return on Equity (ROE)"] = safe_div(net_income, total_stockholders_equity)
    ratios["Return on Capital Employed (ROCE)"] = safe_div(ebit, total_assets - total_current_liabilities)
    ratios["EPS"] = eps
    ratios["P/E Ratio"] = pe
    ratios["Current Ratio"] = safe_div(total_current_assets, total_current_liabilities)
    ratios["Quick Ratio"] = safe_div(total_current_assets - inventory, total_current_liabilities)
    ratios["Cash Ratio"] = safe_div(cash_and_equivalents + short_term_investments, total_current_liabilities)
    ratios["Debt to Equity"] = safe_div(debt_for_ratio, total_stockholders_equity)
    ratios["Debt to Assets"] = safe_div(debt_for_ratio, total_assets)
    ratios["Interest Coverage Ratio"] = safe_div(ebit, interest_expense)
    ratios["Asset Turnover"] = safe_div(revenue, total_assets)
    ratios["Fixed Asset Turnover"] = safe_div(revenue, property_plant_equipment_net)
    working_capital = total_current_assets - total_current_liabilities
    ratios["Working Capital Turnover"] = safe_div(revenue, working_capital)
    ratios["FCF to Net Income"] = safe_div(free_cash_flow, net_income)
    ratios["FCF to Revenue"] = safe_div(free_cash_flow, revenue)
    ratios["CapEx to Operating Cash Flow"] = safe_div(capital_expenditure.abs(), operating_cash_flow)

    return ratios.transpose().multiply(100).round(2)

# === WRITE TO EXCEL ===
def write_excel(file_name, income_df, cash_df, balance_df, ratios_df, forecast_df, anomalies, sentiment, cagr_series):
    """Writes financial data, growth indices, ratios, forecasts, and summary to Excel."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws1 = wb.create_sheet("Financial Model")
    ws2 = wb.create_sheet("Growth Index")
    ws3 = wb.create_sheet("Financial Ratios")
    ws4 = wb.create_sheet("Forecast")
    ws_readme = wb.create_sheet("Read Me")

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    anomaly_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    def write_block(ws, df, title, start_row, percent_format=False, is_ratio_sheet=False):
        if df.empty:
            ws.cell(row=start_row, column=1, value=f"No data for {title}.").font = Font(italic=True)
            return start_row + 2

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(df.columns) + 1)
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color="000080")
        title_cell.alignment = align_left
        start_row += 1

        df_to_write = df.where(pd.notna(df), None)
        for r_idx, row_data in enumerate(dataframe_to_rows(df_to_write, index=True, header=True), start_row):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                if r_idx == start_row or c_idx == 1:
                    cell.font = bold_font
                    if r_idx == start_row:
                        cell.fill = header_fill
                        cell.alignment = align_center
                if isinstance(val, (int, float)):
                    cell.number_format = '0.00%' if percent_format else '0.00' if is_ratio_sheet else '#,##0.0'
                    cell.alignment = align_center
                    if c_idx > 1 and r_idx > start_row and val is not None:
                        row_label = ws.cell(row=r_idx, column=1).value
                        if row_label in anomalies:
                            for anomaly in anomalies[row_label]:
                                if str(val) in anomaly:
                                    cell.fill = anomaly_fill
                else:
                    cell.alignment = align_left

        for col_idx in range(1, df.shape[1] + 2):
            max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(start_row, ws.max_row + 1))
            adjusted_width = max(15, min(max_len + 2, 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        if title.startswith("INCOME STATEMENT"):
            chart = LineChart()
            chart.title = "Revenue & Net Income Trends"
            chart.height = 10
            chart.width = 15
            data = Reference(ws, min_col=2, min_row=start_row + 1, max_col=len(df.columns) + 1, max_row=start_row + 2)
            cats = Reference(ws, min_col=2, min_row=start_row, max_col=len(df.columns) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "A" + str(ws.max_row + 2))
            return ws.max_row + 12
        return ws.max_row + 2

    def write_summary(ws, ticker, income_df, ratios_df, anomalies, sentiment, cagr_series):
        r = 1
        ws.cell(row=r, column=1, value=f"Financial Analysis for {ticker}").font = Font(bold=True, size=16)
        r += 2
        ws.cell(row=r, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S CEST')}").font = Font(italic=True)
        r += 2
        ws.cell(row=r, column=1, value="Data Source: Financial Modeling Prep API").font = Font(italic=True)
        r += 2
        ws.cell(row=r, column=1, value="Key Metrics").font = bold_font
        r += 1
        key_metrics = ["Revenue", "Net Income", "Free Cash Flow", "ROE", "Gross Margin"]
        for metric in key_metrics:
            if metric in income_df.index:
                ws.cell(row=r, column=1, value=metric).font = bold_font
                for c, year in enumerate(income_df.columns, 2):
                    value = income_df.loc[metric, year]
                    ws.cell(row=r, column=c, value=value).number_format = '#,##0.0'
                r += 1
            elif metric in ratios_df.index:
                ws.cell(row=r, column=1, value=metric).font = bold_font
                for c, year in enumerate(ratios_df.columns, 2):
                    value = ratios_df.loc[metric, year]
                    ws.cell(row=r, column=c, value=value).number_format = '0.00%'
                r += 1
        r += 2
        ws.cell(row=r, column=1, value="CAGR (%)").font = bold_font
        r += 1
        for metric, cagr in cagr_series.items():
            if pd.notna(cagr):
                ws.cell(row=r, column=1, value=metric).font = bold_font
                ws.cell(row=r, column=2, value=cagr).number_format = '0.00%'
                r += 1
        r += 2
        ws.cell(row=r, column=1, value="Anomalies").font = bold_font
        r += 1
        for metric, anomaly_list in anomalies.items():
            for anomaly in anomaly_list:
                ws.cell(row=r, column=1, value=f"{metric}: {anomaly}").font = Font(color="FF0000")
                r += 1
        r += 2
        ws.cell(row=r, column=1, value="Sentiment").font = bold_font
        r += 1
        ws.cell(row=r, column=1, value=sentiment)
        r += 2
        for col in range(1, len(income_df.columns) + 2):
            ws.cell(row=7, column=col).border = border
            ws.cell(row=10, column=col).border = border
            ws.cell(row=13, column=col).border = border

    def write_readme(ws):
        ws.cell(row=1, column=1, value="Financial Analysis Report").font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value="This report contains financial data, growth indices, ratios, and forecasts.").font = Font(italic=True)
        ws.cell(row=4, column=1, value="Sheets:").font = bold_font
        ws.cell(row=5, column=1, value="- Summary: Key metrics, anomalies, and sentiment")
        ws.cell(row=6, column=1, value="- Financial Model: Income, cash flow, and balance sheet data")
        ws.cell(row=7, column=1, value="- Growth Index: Growth relative to base year and CAGR")
        ws.cell(row=8, column=1, value="- Financial Ratios: Profitability, liquidity, leverage, and efficiency ratios")
        ws.cell(row=9, column=1, value="- Forecast: ARIMA-based predictions")
        ws.cell(row=11, column=1, value="Notes: All financial data in millions USD. Ratios in percentages.").font = Font(italic=True)

    write_summary(ws_summary, ticker_symbol, income_df, ratios_df, anomalies, sentiment, cagr_series)
    r = 1
    r = write_block(ws1, income_df, "INCOME STATEMENT (Millions USD)", r)
    r = write_block(ws1, cash_df, "CASH FLOW STATEMENT (Millions USD)", r)
    write_block(ws1, balance_df, "BALANCE SHEET (Millions USD)", r)
    r = 1
    growth_income, cagr_income = growth_index(income_df)
    growth_cash, cagr_cash = growth_index(cash_df)
    growth_balance, cagr_balance = growth_index(balance_df)
    r = write_block(ws2, growth_income, "INCOME STATEMENT GROWTH INDEX", r, percent_format=True)
    r = write_block(ws2, growth_cash, "CASH FLOW STATEMENT GROWTH INDEX", r, percent_format=True)
    write_block(ws2, growth_balance, "BALANCE SHEET GROWTH INDEX", r, percent_format=True)
    write_block(ws3, ratios_df, "FINANCIAL RATIOS (%)", 1, is_ratio_sheet=True)
    write_block(ws4, forecast_df, "FORECAST (Millions USD)", 1)
    write_readme(ws_readme)

    wb.save(file_name)
    return file_name

# === EXECUTION PIPELINE ===
def run_financial_analysis(ticker_symbol, years_of_data):
    """Main function to fetch data, calculate, and write to Excel."""
    ticker_symbol = ticker_symbol.strip().upper()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{SAVE_PATH}/Model_{ticker_symbol}_{years_of_data}Y_{timestamp}.xlsx"

    print(f"Fetching data for {ticker_symbol} for {years_of_data} years...")
    income_df = fetch_fmp_data("income-statement", ticker_symbol, years_of_data, API_KEY)
    cash_df = fetch_fmp_data("cash-flow-statement", ticker_symbol, years_of_data, API_KEY)
    balance_df = fetch_fmp_data("balance-sheet-statement", ticker_symbol, years_of_data, API_KEY)
    market_df = fetch_market_data(ticker_symbol, API_KEY)

    if income_df.empty or cash_df.empty or balance_df.empty:
        print(f"Could not fetch complete financial data for {ticker_symbol}.")
        return None

    forecast_metrics = ["Revenue", "Net Income", "Free Cash Flow"]
    forecast_df = pd.DataFrame()
    for metric in forecast_metrics:
        if metric in income_df.index:
            forecast_df[metric] = forecast_metric(income_df.loc[metric])
        elif metric in cash_df.index:
            forecast_df[metric] = forecast_metric(cash_df.loc[metric])

    anomalies = {}
    for df in [income_df, cash_df, balance_df]:
        anomalies.update(detect_anomalies(df))

    sentiment = analyze_sentiment(ticker_symbol)

    ratios_df = calculate_ratios(income_df, balance_df, cf_df, market_df)
    _, cagr_series = growth_index(income_df)

    try:
        written_file = write_excel(output_file, income_df, cash_df, balance_df, ratios_df, forecast_df, anomalies, sentiment, cagr_series)
        print(f"\n✅ Done! Excel file saved to:\n{written_file}")
        return written_file
    except Exception as e:
        print(f"Error writing Excel file: {e}")
        return None
